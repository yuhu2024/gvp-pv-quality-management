#!/usr/bin/env python
"""
Django管理命令 - 从Excel导入用户
用法:
    python manage.py import_users_excel
    python manage.py import_users_excel --file /path/to/users.xlsx
    python manage.py import_users_excel --skip-existing
"""

import os
import sys
import argparse

# 设置Django环境
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'training_system.settings')

import django
django.setup()

from openpyxl import load_workbook
from apps.users.models import User, Department, Role


# Excel列映射配置
COLUMN_MAPPING = {
    'employee_id': '工号',
    'last_name': '姓名',
    'username': '用户名',
    'password': '密码',
    'department_name': '部门',
    'role_name': '角色',
    'phone': '手机号',
}

# 必填字段
REQUIRED_FIELDS = ['username', 'password']

# 默认Excel文件路径
DEFAULT_FILE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'templates_import',
    'user_import_template.xlsx'
)


def get_header_index(headers):
    """
    根据列名获取列索引

    Args:
        headers: Excel表头列表

    Returns:
        dict: 字段名到列索引的映射
    """
    mapping = {}
    for idx, header in enumerate(headers):
        header_str = str(header).strip()
        for field, name in COLUMN_MAPPING.items():
            if header_str == name:
                mapping[field] = idx
                break
    return mapping


def import_users_from_excel(file_path, skip_existing=False):
    """
    从Excel文件导入用户

    Args:
        file_path: Excel文件路径
        skip_existing: 是否跳过已存在的用户

    Returns:
        tuple: (成功数, 失败数, 跳过数, 失败详情列表)
    """
    if not os.path.exists(file_path):
        print(f'错误: 文件不存在 - {file_path}')
        return 0, 0, 0, []

    try:
        workbook = load_workbook(file_path, read_only=True)
    except Exception as e:
        print(f'错误: 无法读取Excel文件 - {e}')
        return 0, 0, 0, []

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if len(rows) < 2:
        print('错误: Excel文件为空或只有表头。')
        return 0, 0, 0, []

    # 解析表头
    headers = rows[0]
    col_map = get_header_index(headers)

    # 检查必填字段
    missing_fields = []
    for field in REQUIRED_FIELDS:
        if field not in col_map:
            missing_fields.append(COLUMN_MAPPING[field])
    if missing_fields:
        print(f'错误: 缺少必填列 - {", ".join(missing_fields)}')
        return 0, 0, 0, []

    print(f'成功解析表头，检测到以下列: {list(col_map.keys())}')
    print(f'开始导入用户数据（共 {len(rows) - 1} 行）...')
    print(f'跳过已存在用户: {"是" if skip_existing else "否"}\n')

    success_count = 0
    fail_count = 0
    skip_count = 0
    fail_details = []

    for row_idx, row in enumerate(rows[1:], start=2):
        # 构建用户数据
        user_data = {}
        for field, col_idx in col_map.items():
            if col_idx < len(row) and row[col_idx] is not None:
                user_data[field] = str(row[col_idx]).strip()

        # 验证必填字段
        if not user_data.get('username') or not user_data.get('password'):
            fail_count += 1
            detail = f'第{row_idx}行: 用户名或密码为空'
            fail_details.append(detail)
            print(f'[失败] {detail}')
            continue

        username = user_data['username']

        try:
            # 校验工号唯一性
            employee_id = user_data.get('employee_id', '')
            if employee_id:
                if User.objects.filter(employee_id=employee_id).exists():
                    if skip_existing:
                        skip_count += 1
                        print(f'[跳过] 第{row_idx}行: 工号 {employee_id} 已存在')
                        continue
                    else:
                        fail_count += 1
                        detail = f'第{row_idx}行: 工号 {employee_id} 已存在'
                        fail_details.append(detail)
                        print(f'[失败] {detail}')
                        continue

            # 校验用户名唯一性
            if User.objects.filter(username=username).exists():
                if skip_existing:
                    skip_count += 1
                    print(f'[跳过] 第{row_idx}行: 用户名 {username} 已存在')
                    continue
                else:
                    fail_count += 1
                    detail = f'第{row_idx}行: 用户名 {username} 已存在'
                    fail_details.append(detail)
                    print(f'[失败] {detail}')
                    continue

            # 获取或创建部门（校验部门存在性）
            department = None
            department_name = user_data.get('department_name', '')
            if department_name:
                department, _ = Department.objects.get_or_create(name=department_name)

            # 获取或创建角色
            role_name = user_data.get('role_name', '')
            role_obj = None
            if role_name:
                role_obj, _ = Role.objects.get_or_create(
                    code=role_name,
                    defaults={'name': role_name}
                )

            # 创建用户
            user = User.objects.create_user(
                username=username,
                password=user_data['password'],
                employee_id=employee_id,
                last_name=user_data.get('last_name', ''),
                phone=user_data.get('phone', ''),
                department=department,
            )

            # 分配角色
            if role_obj:
                user.roles.add(role_obj)

            success_count += 1
            print(f'[成功] 第{row_idx}行: 创建用户 {username}')

        except Exception as e:
            fail_count += 1
            detail = f'第{row_idx}行: {str(e)}'
            fail_details.append(detail)
            print(f'[失败] {detail}')

    workbook.close()

    print(f'\n===== 导入完成 =====')
    print(f'成功: {success_count} 个')
    print(f'跳过: {skip_count} 个')
    print(f'失败: {fail_count} 个')
    if fail_details:
        print(f'\n失败详情（前20条）:')
        for detail in fail_details[:20]:
            print(f'  - {detail}')
        if len(fail_details) > 20:
            print(f'  ... 还有 {len(fail_details) - 20} 条失败记录')

    return success_count, fail_count, skip_count, fail_details


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='从Excel导入用户')
    parser.add_argument(
        '--file', type=str, default=DEFAULT_FILE_PATH,
        help=f'Excel文件路径（默认: {DEFAULT_FILE_PATH}）'
    )
    parser.add_argument(
        '--skip-existing', action='store_true',
        help='跳过已存在的用户（工号或用户名重复时跳过）'
    )

    args = parser.parse_args()

    import_users_from_excel(
        file_path=args.file,
        skip_existing=args.skip_existing,
    )
