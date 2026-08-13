"""
用户与账号管理 - 视图
"""
import secrets
import time

import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout
from django.contrib import messages
from django.core.cache import cache
from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import ensure_csrf_cookie

from .models import User, Department, Role
from .forms import (
    UserLoginForm, UserProfileForm, UserForm, ExcelImportForm
)


def _is_admin_or_manager(user):
    """检查用户是否为管理员或培训管理员"""
    if not user.is_authenticated:
        return False
    return user.is_superuser or (user.role and user.role.code in ('admin', 'training_manager'))


def _get_client_ip(request):
    """获取客户端真实 IP"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def _check_login_rate_limit(request):
    """检查登录尝试频率，返回 True 表示被限制"""
    ip = _get_client_ip(request)
    cache_key = f'login_attempts:{ip}'
    attempts = cache.get(cache_key, [])
    now = time.time()
    # 只保留 60 秒内的尝试
    attempts = [t for t in attempts if now - t < 60]
    if len(attempts) >= 5:
        return True
    attempts.append(now)
    cache.set(cache_key, attempts, 60)
    return False


def _clear_login_rate_limit(request):
    """登录成功后清除限制计数"""
    ip = _get_client_ip(request)
    cache_key = f'login_attempts:{ip}'
    cache.delete(cache_key)


# ==================== 仪表盘 ====================

@login_required
def dashboard_view(request):
    """用户仪表盘 - 显示培训任务概览"""
    user = request.user

    # 统计数据
    pending_course_count = 0
    completed_course_count = 0
    pending_exam_count = 0
    completed_exam_count = 0
    total_learning_hours = 0
    recent_logs = []

    try:
        from plans.models import TaskAssignment
        from logs.models import LearningLog

        # 待完成课程数（任务类型为课程，状态为待开始或进行中）
        pending_course_count = TaskAssignment.objects.filter(
            user=user,
            task__task_type='course',
            status__in=['pending', 'in_progress']
        ).count()

        # 已完成课程数
        completed_course_count = TaskAssignment.objects.filter(
            user=user,
            task__task_type='course',
            status='completed'
        ).count()

        # 待考试数（任务类型为考试，状态为待开始或进行中）
        pending_exam_count = TaskAssignment.objects.filter(
            user=user,
            task__task_type='exam',
            status__in=['pending', 'in_progress']
        ).count()

        # 已完成考试数
        completed_exam_count = TaskAssignment.objects.filter(
            user=user,
            task__task_type='exam',
            status__in=['completed', 'submitted', 'graded']
        ).count()

        # 学习总时长（秒转换为小时）
        from django.db.models import Sum
        total_seconds = LearningLog.objects.filter(user=user).aggregate(
            total=Sum('duration')
        )['total'] or 0
        total_learning_hours = round(total_seconds / 3600, 1)

        # 最近学习记录（最近10条）
        recent_logs = LearningLog.objects.filter(
            user=user
        ).select_related('course', 'exam')[:10]

    except Exception:
        # 如果相关app的模型不存在，使用默认值
        pass

    # 待完成任务列表
    pending_tasks = []
    try:
        from plans.models import TaskAssignment
        pending_tasks = TaskAssignment.objects.filter(
            user=user,
            status__in=['pending', 'in_progress']
        ).select_related('task', 'task__plan', 'task__course', 'task__exam')[:5]
    except Exception:
        pass

    context = {
        'user': user,
        'pending_course_count': pending_course_count,
        'completed_course_count': completed_course_count,
        'pending_exam_count': pending_exam_count,
        'completed_exam_count': completed_exam_count,
        'total_learning_hours': total_learning_hours,
        'recent_logs': recent_logs,
        'pending_tasks': pending_tasks,
        'today': timezone.localdate(),
    }
    return render(request, 'users/dashboard.html', context)


# ==================== 个人资料 ====================

@login_required
def profile_view(request):
    """GET显示个人信息，POST更新个人信息"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '个人资料更新成功！')
            return redirect('users:profile')
        else:
            messages.error(request, '请检查表单中的错误信息。')
    else:
        form = UserProfileForm(instance=request.user)
    return render(request, 'users/profile.html', {'form': form})


# ==================== 用户管理（管理员） ====================

@login_required
def user_list_view(request):
    """管理员查看所有用户列表，支持按部门/角色筛选"""
    if not _is_admin_or_manager(request.user):
        messages.error(request, '您没有权限访问此页面。')
        return redirect('users:dashboard')

    users = User.objects.select_related('department', 'role').all()

    # 筛选参数
    department_id = request.GET.get('department', '')
    role_id = request.GET.get('role', '')
    keyword = request.GET.get('keyword', '')
    is_active = request.GET.get('is_active', '')

    if department_id:
        users = users.filter(department_id=department_id)
    if role_id:
        users = users.filter(role_id=role_id)
    if keyword:
        from django.db.models import Q
        users = users.filter(
            Q(username__icontains=keyword) |
            Q(employee_id__icontains=keyword) |
            Q(last_name__icontains=keyword) |
            Q(first_name__icontains=keyword)
        )
    if is_active in ('true', 'false'):
        users = users.filter(is_active=(is_active == 'true'))

    users = users.order_by('-date_joined')

    # 获取筛选选项
    departments = Department.objects.all()
    roles = Role.objects.all()

    context = {
        'users': users,
        'departments': departments,
        'roles': roles,
        'current_department': department_id,
        'current_role': role_id,
        'current_keyword': keyword,
        'current_is_active': is_active,
    }
    return render(request, 'users/user_list.html', context)


@login_required
def user_create_view(request):
    """管理员创建单个用户"""
    if not _is_admin_or_manager(request.user):
        messages.error(request, '您没有权限访问此页面。')
        return redirect('users:dashboard')

    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'用户 {user.username} 创建成功！')
            return redirect('users:user_list')
        else:
            messages.error(request, '请检查表单中的错误信息。')
    else:
        form = UserForm()

    context = {
        'form': form,
        'page_title': '创建用户',
    }
    return render(request, 'users/user_form.html', context)


@login_required
def user_delete_view(request, pk):
    """管理员删除用户"""
    if not _is_admin_or_manager(request.user):
        messages.error(request, '您没有权限执行此操作。')
        return redirect('users:dashboard')

    user_to_delete = get_object_or_404(User, pk=pk)

    # 不允许删除自己
    if user_to_delete == request.user:
        messages.error(request, '不能删除当前登录的用户。')
        return redirect('users:user_list')

    if request.method == 'POST':
        username = user_to_delete.username
        user_to_delete.delete()
        messages.success(request, f'用户 {username} 已成功删除。')
        return redirect('users:user_list')

    context = {
        'user_obj': user_to_delete,
    }
    return render(request, 'users/user_confirm_delete.html', context)


@login_required
def batch_import_view(request):
    """Excel批量导入用户页面"""
    if not _is_admin_or_manager(request.user):
        messages.error(request, '您没有权限访问此页面。')
        return redirect('users:dashboard')

    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            result = _process_excel_import(excel_file, request)
            if result['success']:
                messages.success(request, result['message'])
            else:
                messages.warning(request, result['message'])
            return redirect('users:batch_import')
        else:
            messages.error(request, '请检查表单中的错误信息。')
    else:
        form = ExcelImportForm()

    context = {
        'form': form,
    }
    return render(request, 'users/batch_import.html', context)


def _process_excel_import(excel_file, request):
    """
    处理Excel文件导入用户
    Excel列要求：工号、用户名、姓名（姓）、名字（名）、邮箱、手机号、部门编码、角色编码、职位
    """
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        ws = wb.active

        # 读取表头（第一行）
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else '')

        # 建立列名到索引的映射
        col_map = {}
        for idx, header in enumerate(headers):
            col_map[header] = idx

        # 验证必要列
        required_columns = ['工号', '用户名']
        missing = [col for col in required_columns if col not in col_map]
        if missing:
            return {
                'success': False,
                'message': f'Excel缺少必要列：{", ".join(missing)}。请确保表头包含：工号、用户名、姓名（可选）、名字（可选）、邮箱（可选）、手机号（可选）、部门编码（可选）、角色编码（可选）、职位（可选）'
            }

        created_count = 0
        updated_count = 0
        error_rows = []
        row_num = 0

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not row or all(cell is None for cell in row):
                    continue

                def get_col(name):
                    idx = col_map.get(name)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val is not None else ''
                    return ''

                employee_id = get_col('工号')
                username = get_col('用户名')
                last_name = get_col('姓名') or get_col('姓')
                first_name = get_col('名字') or get_col('名')
                email = get_col('邮箱')
                phone = get_col('手机号')
                department_code = get_col('部门编码')
                role_code = get_col('角色编码')
                position = get_col('职位')

                if not employee_id or not username:
                    error_rows.append(f'第{row_num + 1}行：工号或用户名为空，跳过')
                    continue

                # 查找或创建部门
                department = None
                if department_code:
                    try:
                        department = Department.objects.get(code=department_code)
                    except Department.DoesNotExist:
                        error_rows.append(f'第{row_num + 1}行：部门编码"{department_code}"不存在，跳过')
                        continue

                # 查找或创建角色
                role = None
                if role_code:
                    try:
                        role = Role.objects.get(code=role_code)
                    except Role.DoesNotExist:
                        error_rows.append(f'第{row_num + 1}行：角色编码"{role_code}"不存在，跳过')
                        continue

                # 查找是否已存在该工号或用户名的用户
                existing_user = User.objects.filter(employee_id=employee_id).first()
                existing_by_username = User.objects.filter(username=username).first()

                if existing_user:
                    # 更新已有用户
                    existing_user.last_name = last_name
                    existing_user.first_name = first_name
                    existing_user.email = email
                    existing_user.phone = phone
                    existing_user.department = department
                    existing_user.role = role
                    existing_user.position = position
                    existing_user.save()
                    updated_count += 1
                elif existing_by_username:
                    error_rows.append(f'第{row_num + 1}行：用户名"{username}"已存在但工号不同，跳过')
                    continue
                else:
                    # 创建新用户，使用随机密码（不再使用工号作为密码）
                    temp_password = secrets.token_urlsafe(12)
                    new_user = User(
                        username=username,
                        employee_id=employee_id,
                        last_name=last_name,
                        first_name=first_name,
                        email=email,
                        phone=phone,
                        department=department,
                        role=role,
                        position=position,
                        is_active=True,
                    )
                    new_user.set_password(temp_password)
                    new_user.save()
                    # 记录临时密码到日志（管理员可查看）
                    error_rows.append(
                        f'第{row_num + 1}行：用户 {username} 已创建，临时密码：{temp_password}（请通知用户尽快登录修改）'
                    )
                    created_count += 1

        wb.close()

        msg_parts = [f'导入完成：成功创建 {created_count} 个用户，更新 {updated_count} 个用户。']
        if error_rows:
            msg_parts.append(f'失败 {len(error_rows)} 行：{"; ".join(error_rows[:5])}')
            if len(error_rows) > 5:
                msg_parts.append(f'（还有 {len(error_rows) - 5} 条错误...）')

        return {
            'success': True,
            'message': ''.join(msg_parts)
        }

    except Exception as e:
        return {
            'success': False,
            'message': f'导入失败：{str(e)}'
        }


# ==================== 登录/登出 ====================

@ensure_csrf_cookie
def login_view(request):
    """用户登录（带 CSRF 保护和速率限制）"""
    if request.user.is_authenticated:
        return redirect('users:dashboard')
    if request.method == 'POST':
        # 检查登录频率限制
        if _check_login_rate_limit(request):
            messages.error(request, '登录尝试过于频繁，请 60 秒后再试。')
            form = UserLoginForm()
            return render(request, 'registration/login.html', {'form': form})

        form = UserLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            _clear_login_rate_limit(request)
            messages.success(request, f'欢迎回来，{user}！')
            # 检查是否需要强制修改密码
            if getattr(user, 'force_password_change', False):
                messages.warning(request, '请先修改您的初始密码后再继续操作。')
                return redirect('users:change_password')
            # 验证 next_url 防止开放重定向
            next_url = request.GET.get('next') or request.POST.get('next') or '/'
            if not url_has_allowed_host_and_scheme(
                next_url, allowed_hosts={request.get_host()}
            ):
                next_url = '/'
            return redirect(next_url)
    else:
        form = UserLoginForm()
    return render(request, 'registration/login.html', {'form': form})


def logout_view(request):
    """用户登出"""
    logout(request)
    messages.info(request, '您已成功退出登录。')
    return redirect('users:login')


@login_required
def change_password_view(request):
    """修改密码（支持强制修改）"""
    if request.method == 'POST':
        old_password = request.POST.get('old_password', '')
        new_password1 = request.POST.get('new_password1', '')
        new_password2 = request.POST.get('new_password2', '')

        user = request.user

        # 验证旧密码（非强制修改时才验证）
        if not getattr(user, 'force_password_change', False):
            if not user.check_password(old_password):
                messages.error(request, '旧密码不正确。')
                return render(request, 'registration/change_password.html')

        # 验证新密码
        if not new_password1:
            messages.error(request, '请输入新密码。')
            return render(request, 'registration/change_password.html')

        if len(new_password1) < 8:
            messages.error(request, '密码长度不能少于8位。')
            return render(request, 'registration/change_password.html')

        if new_password1.isdigit():
            messages.error(request, '密码不能全为数字。')
            return render(request, 'registration/change_password.html')

        if new_password1 != new_password2:
            messages.error(request, '两次输入的密码不一致。')
            return render(request, 'registration/change_password.html')

        # 检查密码与用户名相似度
        from django.contrib.auth.password_validation import UserAttributeSimilarityValidator
        validator = UserAttributeSimilarityValidator()
        try:
            validator.validate(new_password1, user)
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'registration/change_password.html')

        # 设置新密码
        user.set_password(new_password1)
        user.force_password_change = False
        user.save()

        # 重新登录（因为 set_password 会使用户的 session 失效）
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, user)

        messages.success(request, '密码修改成功！')
        return redirect('users:dashboard')

    return render(request, 'registration/change_password.html')
