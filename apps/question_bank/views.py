"""
题库管理 - 视图
"""
import json
import io
from collections import Counter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.views.decorators.http import require_POST

from .models import KnowledgePoint, QuestionBank


def is_admin(user):
    return user.is_staff or user.is_superuser


# ==================== 题库列表与筛选 ====================

@login_required
@user_passes_test(is_admin)
def question_list_view(request):
    """题库列表 - 支持多维度筛选、搜索"""
    questions = QuestionBank.objects.select_related('created_by').prefetch_related('knowledge_points')

    # 筛选条件
    q_type = request.GET.get('type', '')
    difficulty = request.GET.get('difficulty', '')
    knowledge_point = request.GET.get('kp', '')
    tag = request.GET.get('tag', '')
    search = request.GET.get('search', '')
    active_only = request.GET.get('active', '') != 'all'

    if q_type:
        questions = questions.filter(question_type=q_type)
    if difficulty:
        questions = questions.filter(difficulty=difficulty)
    if knowledge_point:
        questions = questions.filter(knowledge_points__id=knowledge_point)
    if tag:
        questions = questions.filter(tags__icontains=tag)
    if search:
        questions = questions.filter(
            Q(question_text__icontains=search) | Q(tags__icontains=search)
        )
    if active_only:
        questions = questions.filter(is_active=True)

    # 统计
    stats = QuestionBank.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(is_active=True)),
        easy=Count('id', filter=Q(difficulty='easy')),
        medium=Count('id', filter=Q(difficulty='medium')),
        hard=Count('id', filter=Q(difficulty='hard')),
    )

    # 所有标签
    all_tags = set()
    for q in QuestionBank.objects.exclude(tags='').values_list('tags', flat=True):
        for t in q.split(','):
            t = t.strip()
            if t:
                all_tags.add(t)

    return render(request, 'question_bank/list.html', {
        'questions': questions,
        'knowledge_points': KnowledgePoint.objects.all(),
        'stats': stats,
        'all_tags': sorted(all_tags),
        'filters': {
            'type': q_type, 'difficulty': difficulty,
            'kp': knowledge_point, 'tag': tag,
            'search': search, 'active': '' if active_only else 'all',
        },
        'type_choices': QuestionBank.QUESTION_TYPE_CHOICES,
        'difficulty_choices': QuestionBank.DIFFICULTY_CHOICES,
    })


# ==================== 题目 CRUD ====================

@login_required
@user_passes_test(is_admin)
def question_create_view(request):
    """创建题目"""
    if request.method == 'POST':
        data = request.POST
        question = QuestionBank.objects.create(
            question_text=data.get('question_text', ''),
            question_type=data.get('question_type', 'single_choice'),
            option_a=data.get('option_a', ''),
            option_b=data.get('option_b', ''),
            option_c=data.get('option_c', ''),
            option_d=data.get('option_d', ''),
            correct_answer=data.get('correct_answer', ''),
            analysis=data.get('analysis', ''),
            score=int(data.get('score', 5)),
            difficulty=data.get('difficulty', 'medium'),
            tags=data.get('tags', ''),
            created_by=request.user,
        )
        # 关联知识点
        kp_ids = request.POST.getlist('knowledge_points')
        if kp_ids:
            question.knowledge_points.set(kp_ids)

        messages.success(request, '题目已添加到题库。')
        return redirect('question_bank:list')

    return render(request, 'question_bank/form.html', {
        'knowledge_points': KnowledgePoint.objects.all(),
        'type_choices': QuestionBank.QUESTION_TYPE_CHOICES,
        'difficulty_choices': QuestionBank.DIFFICULTY_CHOICES,
    })


@login_required
@user_passes_test(is_admin)
def question_edit_view(request, pk):
    """编辑题目"""
    question = get_object_or_404(QuestionBank, pk=pk)
    if request.method == 'POST':
        data = request.POST
        question.question_text = data.get('question_text', '')
        question.question_type = data.get('question_type', 'single_choice')
        question.option_a = data.get('option_a', '')
        question.option_b = data.get('option_b', '')
        question.option_c = data.get('option_c', '')
        question.option_d = data.get('option_d', '')
        question.correct_answer = data.get('correct_answer', '')
        question.analysis = data.get('analysis', '')
        question.score = int(data.get('score', 5))
        question.difficulty = data.get('difficulty', 'medium')
        question.tags = data.get('tags', '')
        question.save()

        kp_ids = request.POST.getlist('knowledge_points')
        question.knowledge_points.set(kp_ids)

        messages.success(request, '题目已更新。')
        return redirect('question_bank:list')

    return render(request, 'question_bank/form.html', {
        'question': question,
        'knowledge_points': KnowledgePoint.objects.all(),
        'type_choices': QuestionBank.QUESTION_TYPE_CHOICES,
        'difficulty_choices': QuestionBank.DIFFICULTY_CHOICES,
    })


@login_required
@user_passes_test(is_admin)
@require_POST
def question_delete_view(request, pk):
    """删除题目"""
    question = get_object_or_404(QuestionBank, pk=pk)
    question.delete()
    messages.success(request, '题目已删除。')
    return redirect('question_bank:list')


@login_required
@user_passes_test(is_admin)
@require_POST
def question_batch_delete_view(request):
    """批量删除题目"""
    ids = request.POST.getlist('ids')
    if ids:
        QuestionBank.objects.filter(id__in=ids).delete()
        messages.success(request, f'已删除 {len(ids)} 道题目。')
    return redirect('question_bank:list')


# ==================== 批量导入导出 ====================

@login_required
@user_passes_test(is_admin)
def question_export_view(request):
    """导出题库为 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '题库'

    # 表头
    headers = ['题目内容', '题目类型', '选项A', '选项B', '选项C', '选项D',
               '正确答案', '答案解析', '分值', '难度', '知识点', '标签']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    type_map = dict(QuestionBank.QUESTION_TYPE_CHOICES)
    difficulty_map = dict(QuestionBank.DIFFICULTY_CHOICES)

    questions = QuestionBank.objects.select_related().prefetch_related('knowledge_points')
    for row, q in enumerate(questions, 2):
        ws.cell(row=row, column=1, value=q.question_text)
        ws.cell(row=row, column=2, value=type_map.get(q.question_type, q.question_type))
        ws.cell(row=row, column=3, value=q.option_a)
        ws.cell(row=row, column=4, value=q.option_b)
        ws.cell(row=row, column=5, value=q.option_c)
        ws.cell(row=row, column=6, value=q.option_d)
        ws.cell(row=row, column=7, value=q.correct_answer)
        ws.cell(row=row, column=8, value=q.analysis)
        ws.cell(row=row, column=9, value=q.score)
        ws.cell(row=row, column=10, value=difficulty_map.get(q.difficulty, q.difficulty))
        ws.cell(row=row, column=11, value=', '.join(kp.name for kp in q.knowledge_points.all()))
        ws.cell(row=row, column=12, value=q.tags)

    # 列宽
    widths = [45, 12, 25, 25, 25, 25, 20, 30, 8, 8, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=question_bank_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def question_import_view(request):
    """从 Excel 批量导入题库"""
    import openpyxl

    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, '请选择要导入的 Excel 文件。')
            return redirect('question_bank:list')

        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            ws = wb.active

            type_reverse = {v: k for k, v in QuestionBank.QUESTION_TYPE_CHOICES}
            difficulty_reverse = {v: k for k, v in QuestionBank.DIFFICULTY_CHOICES}

            created = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue

                question_text = str(row[0]).strip() if row[0] else ''
                if not question_text:
                    skipped += 1
                    continue

                q_type = type_reverse.get(str(row[1]).strip() if row[1] else '单选题', 'single_choice')
                difficulty = difficulty_reverse.get(str(row[9]).strip() if len(row) > 9 and row[9] else '中等', 'medium')

                try:
                    score = int(row[8]) if len(row) > 8 and row[8] else 5
                except (ValueError, TypeError):
                    score = 5

                QuestionBank.objects.create(
                    question_text=question_text,
                    question_type=q_type,
                    option_a=str(row[2]).strip() if len(row) > 2 and row[2] else '',
                    option_b=str(row[3]).strip() if len(row) > 3 and row[3] else '',
                    option_c=str(row[4]).strip() if len(row) > 4 and row[4] else '',
                    option_d=str(row[5]).strip() if len(row) > 5 and row[5] else '',
                    correct_answer=str(row[6]).strip() if len(row) > 6 and row[6] else '',
                    analysis=str(row[7]).strip() if len(row) > 7 and row[7] else '',
                    score=score,
                    difficulty=difficulty,
                    tags=str(row[11]).strip() if len(row) > 11 and row[11] else '',
                    created_by=request.user,
                )
                created += 1

            messages.success(request, f'成功导入 {created} 道题目，跳过 {skipped} 行。')
        except Exception as e:
            messages.error(request, f'导入失败：{e}')

        return redirect('question_bank:list')

    return render(request, 'question_bank/import.html')


# ==================== 知识点管理 ====================

@login_required
@user_passes_test(is_admin)
def knowledge_point_list_view(request):
    """知识点列表"""
    kps = KnowledgePoint.objects.select_related('parent').all()
    return render(request, 'question_bank/knowledge_points.html', {
        'knowledge_points': kps,
    })


@login_required
@user_passes_test(is_admin)
def knowledge_point_create_view(request):
    """创建知识点"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        parent_id = request.POST.get('parent')
        description = request.POST.get('description', '').strip()
        order = int(request.POST.get('order', 0))

        if name:
            KnowledgePoint.objects.create(
                name=name,
                parent_id=parent_id if parent_id else None,
                description=description,
                order=order,
            )
            messages.success(request, f'知识点 "{name}" 已创建。')
        return redirect('question_bank:knowledge_points')

    return render(request, 'question_bank/kp_form.html', {
        'parents': KnowledgePoint.objects.filter(parent__isnull=True),
    })


# ==================== 统计分析 ====================

@login_required
@user_passes_test(is_admin)
def question_stats_view(request):
    """题库统计"""
    stats = {
        'total': QuestionBank.objects.count(),
        'active': QuestionBank.objects.filter(is_active=True).count(),
        'by_type': dict(QuestionBank.objects.values_list('question_type').annotate(c=Count('id')).order_by()),
        'by_difficulty': dict(QuestionBank.objects.values_list('difficulty').annotate(c=Count('id')).order_by()),
        'by_kp': list(KnowledgePoint.objects.annotate(c=Count('questions')).filter(c__gt=0).values('name', 'c').order_by('-c')[:10]),
        'top_tags': Counter(),
    }
    from collections import Counter as Ctr
    tag_counter = Ctr()
    for tags in QuestionBank.objects.exclude(tags='').values_list('tags', flat=True):
        for t in tags.split(','):
            t = t.strip()
            if t:
                tag_counter[t] += 1
    stats['top_tags'] = tag_counter.most_common(15)

    return render(request, 'question_bank/stats.html', {
        'stats': stats,
        'type_choices': dict(QuestionBank.QUESTION_TYPE_CHOICES),
        'difficulty_choices': dict(QuestionBank.DIFFICULTY_CHOICES),
    })