"""
学习痕迹记录 - 视图
"""
import datetime

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.views.generic import ListView
from django.utils.decorators import method_decorator
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Q, F, Max
from django.utils import timezone

from .models import LearningLog, OperationLog
from apps.users.models import User
from apps.plans.models import TaskAssignment
from apps.exams.models import ExamAttempt


def is_admin(user):
    """检查用户是否为管理员"""
    return user.is_staff or user.is_superuser


@method_decorator(login_required, name='dispatch')
class LearningLogView(ListView):
    """学习痕迹记录列表"""
    model = LearningLog
    template_name = 'logs/learning_log.html'
    context_object_name = 'logs'
    paginate_by = 20

    def get_queryset(self):
        # 管理员可查看所有用户记录，普通用户只能查看自己的记录
        if self.request.user.is_staff or self.request.user.is_superuser:
            queryset = LearningLog.objects.all()
        else:
            queryset = LearningLog.objects.filter(user=self.request.user)

        # 按用户筛选（仅管理员）
        user_id = self.request.GET.get('user')
        if user_id and (self.request.user.is_staff or self.request.user.is_superuser):
            queryset = queryset.filter(user_id=user_id)

        # 按操作类型筛选
        action_type = self.request.GET.get('action_type')
        if action_type:
            queryset = queryset.filter(action_type=action_type)

        # 按时间范围筛选
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            try:
                start_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__date__gte=start_dt.date())
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__date__lte=end_dt.date())
            except ValueError:
                pass

        return queryset.select_related('user', 'course', 'exam', 'material')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_choices'] = LearningLog.ACTION_CHOICES
        context['current_action_type'] = self.request.GET.get('action_type', '')
        context['current_user'] = self.request.GET.get('user', '')
        context['current_start_date'] = self.request.GET.get('start_date', '')
        context['current_end_date'] = self.request.GET.get('end_date', '')
        context['is_admin'] = self.request.user.is_staff or self.request.user.is_superuser

        # 如果是管理员，提供用户列表供筛选
        if context['is_admin']:
            context['users'] = User.objects.all().order_by('username')

        return context


@method_decorator(login_required, name='dispatch')
class OperationLogView(ListView):
    """操作日志列表（仅管理员）"""
    model = OperationLog
    template_name = 'logs/operation_log.html'
    context_object_name = 'logs'
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        # 仅管理员可访问
        if not request.user.is_staff and not request.user.is_superuser:
            from django.contrib import messages
            messages.error(request, '您没有权限查看操作日志。')
            from django.shortcuts import redirect
            return redirect('users:dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = OperationLog.objects.all()

        # 按用户筛选
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        # 按操作类型筛选
        action = self.request.GET.get('action')
        if action:
            queryset = queryset.filter(action=action)

        # 按时间范围筛选
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            try:
                start_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__date__gte=start_dt.date())
            except ValueError:
                pass
        if end_date:
            try:
                end_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__date__lte=end_dt.date())
            except ValueError:
                pass

        return queryset.select_related('user')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_choices'] = OperationLog.ACTION_CHOICES
        context['target_type_choices'] = OperationLog.TARGET_TYPE_CHOICES
        context['current_action'] = self.request.GET.get('action', '')
        context['current_user'] = self.request.GET.get('user', '')
        context['current_start_date'] = self.request.GET.get('start_date', '')
        context['current_end_date'] = self.request.GET.get('end_date', '')
        context['users'] = User.objects.all().order_by('username')
        return context


@login_required
def user_statistics_view(request, user_id):
    """用户培训统计"""
    if not request.user.is_staff and not request.user.is_superuser:
        from django.contrib import messages
        messages.error(request, '您没有权限查看用户统计。')
        from django.shortcuts import redirect
        return redirect('users:dashboard')

    target_user = get_object_or_404(User, pk=user_id)

    # 课程学习进度
    course_logs = LearningLog.objects.filter(
        user=target_user,
        action_type='view_course'
    ).values('course__id', 'course__title').annotate(
        view_count=Count('id'),
        total_duration=Sum('duration')
    ).order_by('course__title')

    # 考试成绩
    exam_attempts = ExamAttempt.objects.filter(
        user=target_user,
        status__in=['completed', 'timeout']
    ).select_related('exam').order_by('-start_time')

    # 任务完成率
    total_tasks = TaskAssignment.objects.filter(user=target_user).count()
    completed_tasks = TaskAssignment.objects.filter(
        user=target_user, status='completed'
    ).count()
    task_completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

    # 总学习时长
    total_duration = LearningLog.objects.filter(
        user=target_user
    ).aggregate(total=Sum('duration'))['total'] or 0

    # 学习记录总数
    total_logs = LearningLog.objects.filter(user=target_user).count()

    # 各操作类型统计
    action_stats = LearningLog.objects.filter(
        user=target_user
    ).values('action_type').annotate(
        count=Count('id')
    ).order_by('-count')

    context = {
        'target_user': target_user,
        'course_logs': course_logs,
        'exam_attempts': exam_attempts,
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'task_completion_rate': round(task_completion_rate, 1),
        'total_duration': total_duration,
        'total_logs': total_logs,
        'action_stats': action_stats,
    }
    return render(request, 'logs/user_statistics.html', context)


@login_required
def export_learning_log_view(request):
    """导出学习记录为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        from django.contrib import messages
        messages.error(request, '导出功能需要安装 openpyxl 库。')
        from django.shortcuts import redirect
        return redirect('logs:learning_log')

    # 获取查询参数（与列表视图一致的筛选逻辑）
    if request.user.is_staff or request.user.is_superuser:
        queryset = LearningLog.objects.all()
    else:
        queryset = LearningLog.objects.filter(user=request.user)

    user_id = request.GET.get('user')
    if user_id and (request.user.is_staff or request.user.is_superuser):
        queryset = queryset.filter(user_id=user_id)

    action_type = request.GET.get('action_type')
    if action_type:
        queryset = queryset.filter(action_type=action_type)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        try:
            start_dt = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            queryset = queryset.filter(created_at__date__gte=start_dt.date())
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.datetime.strptime(end_date, '%Y-%m-%d')
            queryset = queryset.filter(created_at__date__lte=end_dt.date())
        except ValueError:
            pass

    queryset = queryset.select_related('user', 'course', 'exam', 'material').order_by('-created_at')

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学习记录'

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = ['序号', '用户', '操作类型', '课程', '考试', '资料', '详情',
               'IP地址', '停留时长(秒)', '记录时间']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 写入数据
    for row_idx, log in enumerate(queryset, start=2):
        data = [
            row_idx - 1,
            str(log.user),
            log.get_action_type_display(),
            log.course.title if log.course else '',
            log.exam.title if log.exam else '',
            log.material.title if log.material else '',
            log.detail,
            log.ip_address or '',
            log.duration or '',
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
        ]
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 调整列宽
    column_widths = [6, 15, 12, 20, 20, 20, 30, 15, 12, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=learning_logs.xlsx'
    wb.save(response)
    return response


# ==================== 管理员数据统计面板 ====================

@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    """管理员数据统计面板"""
    from apps.users.models import Department
    from apps.courses.models import Course, CourseProgress
    from apps.exams.models import Exam
    from apps.plans.models import TrainingPlan
    from datetime import timedelta

    today = timezone.localdate()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)

    # 基础统计
    total_users = User.objects.filter(is_active=True).count()
    total_courses = Course.objects.filter(status='published').count()
    total_exams = Exam.objects.filter(is_published=True).count()
    total_plans = TrainingPlan.objects.count()

    # 近7天活跃用户
    active_users_7d = LearningLog.objects.filter(
        created_at__date__gte=last_7_days
    ).values('user').distinct().count()

    # 近30天活跃用户
    active_users_30d = LearningLog.objects.filter(
        created_at__date__gte=last_30_days
    ).values('user').distinct().count()

    # 课程完成情况
    completed_count = CourseProgress.objects.filter(is_completed=True).count()
    total_enrolled = CourseProgress.objects.count()
    completion_rate = (completed_count / total_enrolled * 100) if total_enrolled > 0 else 0

    # 考试通过率
    exam_attempts = ExamAttempt.objects.filter(status__in=['completed', 'timeout'])
    passed_attempts = exam_attempts.filter(is_passed=True).count()
    total_attempts = exam_attempts.count()
    pass_rate = (passed_attempts / total_attempts * 100) if total_attempts > 0 else 0

    # 平均考试成绩
    avg_score = exam_attempts.aggregate(avg=Avg('score'))['avg'] or 0

    # 部门统计
    dept_stats = Department.objects.annotate(
        user_count=Count('user', filter=Q(user__is_active=True)),
        avg_score=Avg('user__exam_attempts__score'),
    ).order_by('-user_count')[:10]

    # 课程热度TOP10
    top_courses = Course.objects.annotate(
        enrollment_count=Count('user_progress'),
        completed_count=Count('user_progress', filter=Q(user_progress__is_completed=True)),
    ).filter(status='published').order_by('-enrollment_count')[:10]

    # 近7天每日学习人数趋势
    daily_active = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        count = LearningLog.objects.filter(
            created_at__date=date
        ).values('user').distinct().count()
        daily_active.append({'date': date.strftime('%m-%d'), 'count': count})

    # 任务完成统计
    task_stats = TaskAssignment.objects.aggregate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        in_progress=Count('id', filter=Q(status='in_progress')),
        overdue=Count('id', filter=Q(status='overdue')),
    )

    context = {
        'total_users': total_users,
        'total_courses': total_courses,
        'total_exams': total_exams,
        'total_plans': total_plans,
        'active_users_7d': active_users_7d,
        'active_users_30d': active_users_30d,
        'completion_rate': round(completion_rate, 1),
        'pass_rate': round(pass_rate, 1),
        'avg_score': round(avg_score, 1),
        'completed_count': completed_count,
        'total_enrolled': total_enrolled,
        'dept_stats': dept_stats,
        'top_courses': top_courses,
        'daily_active': daily_active,
        'task_stats': task_stats,
        'today': today,
    }
    return render(request, 'logs/admin_dashboard.html', context)


# ==================== 报表导出视图 ====================

@login_required
@user_passes_test(is_admin)
def export_exam_scores_view(request):
    """导出考试成绩为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messages.error(request, '导出功能需要安装 openpyxl 库。')
        return redirect('logs:admin_dashboard')

    from apps.exams.models import Exam

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '考试成绩'

    headers = ['序号', '用户名', '姓名', '部门', '考试名称', '分数', '总分', '是否及格', '状态', '开始时间', '结束时间']
    ws.append(headers)

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    attempts = ExamAttempt.objects.select_related('user', 'user__department', 'exam').order_by('-start_time')
    for i, attempt in enumerate(attempts, 1):
        ws.append([
            i, attempt.user.username, attempt.user.get_full_name() or attempt.user.username,
            str(attempt.user.department) if attempt.user.department else '',
            attempt.exam.title, attempt.score or 0, attempt.exam.total_score,
            '是' if attempt.is_passed else '否', attempt.get_status_display(),
            str(attempt.start_time) if attempt.start_time else '',
            str(attempt.end_time) if attempt.end_time else ''
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=exam_scores.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_admin)
def export_training_report_view(request):
    """导出培训完成情况"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messages.error(request, '导出功能需要安装 openpyxl 库。')
        return redirect('logs:admin_dashboard')

    from apps.courses.models import Course, CourseProgress
    from apps.plans.models import TrainingPlan

    wb = openpyxl.Workbook()

    # Sheet 1: 用户培训概览
    ws1 = wb.active
    ws1.title = '用户培训概览'
    headers = ['序号', '工号', '姓名', '部门', '角色', '已学课程', '已完成课程', '完成率(%)',
               '考试次数', '最高均分', '任务总数', '已完成任务', '总学习时长(小时)']
    ws1.append(headers)

    users = User.objects.filter(is_active=True).select_related('department', 'role')
    for i, user in enumerate(users, 1):
        enrolled = CourseProgress.objects.filter(user=user).count()
        completed = CourseProgress.objects.filter(user=user, is_completed=True).count()
        comp_rate = (completed / enrolled * 100) if enrolled > 0 else 0

        exam_atts = ExamAttempt.objects.filter(
            user=user, status__in=['completed', 'timeout']
        )
        exam_count = exam_atts.count()
        max_score = exam_atts.aggregate(max_s=Max('score'))['max_s'] or 0

        total_tasks = TaskAssignment.objects.filter(user=user).count()
        done_tasks = TaskAssignment.objects.filter(user=user, status='completed').count()

        total_dur = LearningLog.objects.filter(user=user).aggregate(
            total=Sum('duration')
        )['total'] or 0
        total_hours = round(total_dur / 3600, 1)

        ws1.append([
            i, user.employee_id or '', user.get_full_name() or user.username,
            str(user.department) if user.department else '',
            str(user.role) if user.role else '',
            enrolled, completed, round(comp_rate, 1),
            exam_count, max_score,
            total_tasks, done_tasks, total_hours
        ])

    # Sheet 2: 培训计划完成情况
    ws2 = wb.create_sheet('计划完成情况')
    headers2 = ['序号', '计划名称', '状态', '创建时间', '开始时间', '结束时间', '任务数', '已完成', '完成率(%)']
    ws2.append(headers2)

    plans = TrainingPlan.objects.all().order_by('-created_at')
    for i, plan in enumerate(plans, 1):
        total_tasks = plan.tasks.count()
        done_tasks = TaskAssignment.objects.filter(
            task__plan=plan, status='completed'
        ).count()
        plan_rate = (done_tasks / total_tasks * 100) if total_tasks > 0 else 0

        ws2.append([
            i, plan.title, plan.get_status_display(),
            str(plan.created_at.strftime('%Y-%m-%d') if plan.created_at else ''),
            str(plan.start_date.strftime('%Y-%m-%d') if hasattr(plan, 'start_date') and plan.start_date else ''),
            str(plan.end_date.strftime('%Y-%m-%d') if hasattr(plan, 'end_date') and plan.end_date else ''),
            total_tasks, done_tasks, round(plan_rate, 1)
        ])

    # 样式
    for ws in [ws1, ws2]:
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=training_report.xlsx'
    wb.save(response)
    return response
